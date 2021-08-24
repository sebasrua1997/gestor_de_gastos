import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from tkinter import *


wb = load_workbook('control_gastos.xlsx')
pestaña = wb['Hoja1']
print("                 BIENVENIDO A TU CONTROLADOR DE GASTOS\n")

while True:
    print("Ingrese si es una entrada o una salida: ")
    tipo_dato = input()

    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row

    fila_totales = max_fila
    col_total_ingresos = 2
    col_total_egresos = 4

    if tipo_dato == "entrada":
        print("Describa el ingreso: ")
        descripcion_ingreso = input()
        print("¿Cuanto dinero ingreso?: ")
        ingreso = float(input())
        pestaña['A'+str(fila_totales+1)] = descripcion_ingreso
        pestaña['B'+str(fila_totales+1)] = ingreso
        pestaña['C'+str(fila_totales+1)] = "-----"
        pestaña['D'+str(fila_totales+1)] = 0

        wb.save('control_gastos.xlsx')

    elif tipo_dato == "salida":

        print("Describa la salida: ")
        descripcion_salida = input()
        print("¿Cuanto dinero salio?: ")
        salida = float(input())
        pestaña['A'+str(fila_totales+1)] = "-----"
        pestaña['B'+str(fila_totales+1)] = 0
        pestaña['C'+str(fila_totales+1)] = descripcion_salida
        pestaña['D'+str(fila_totales+1)] = salida

        wb.save('control_gastos.xlsx')

    print("Desea continuar? y/n")
    continuacion = input()
    if continuacion == "n":
        print("Finalizando actualizacion...\n\n")
        break
    else:
        print("El programa continuara!\n\n")

max_fila = wb.active.max_row
pestaña['G2'] = '=SUM(B2:B100)' #Total_Ingresos
pestaña['H2'] = '=SUM(D2:D100)' #Total suma_salidas
wb.save('control_gastos.xlsx')

#Analisis de entradas respecto a las salidas

#Entradas
suma_entradas = []
sheet = wb.active
for row in sheet.iter_rows(min_row=2, min_col=2, max_row=max_fila, max_col=2):
    for cell in row:
        suma_entradas.append(float(cell.value))
print("Sus ingresos hasta el momento son: ",sum(suma_entradas))

#salidas
suma_salidas = []
sheet = wb.active
for row in sheet.iter_rows(min_row=2, min_col=4, max_row=max_fila, max_col=4):
    for cell in row:
        suma_salidas.append(float(cell.value))
print("Sus salidas hasta el momento son: ",sum(suma_salidas))
porcentaje = (sum(suma_salidas)/sum(suma_entradas))*100
print("Sus gastos en este momento corresponden al {}% de sus ingresos".format(porcentaje))

#Interfaz Grafica
